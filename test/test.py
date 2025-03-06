import os
import subprocess
import re
import pandas as pd
from openpyxl import load_workbook

query_dir = './insert/query_graph/'
output_file = 'test.xlsx'

# 辅助函数：从文件名中提取数字
def extract_number(filename):
    match = re.search(r'\d+', filename)
    return int(match.group()) if match else float('inf')

# 获取 query_dir 下所有文件，并按文件名中数字排序
filenames = sorted(
    [f for f in os.listdir(query_dir) if os.path.isfile(os.path.join(query_dir, f))],
    key=extract_number
)

results = []

for filename in filenames:
    file_path = os.path.join(query_dir, filename)
    
    command = [
        '../build/streaming/HybridFlow.out',
        '-d', './insert/data_graph/data.graph',
        '-q', file_path,
        '-u', './insert/data_graph/insertion.graph',
        '-num', '4294967296',
        '-time_limit', '3600'
    ]

    try:
        output = subprocess.run(
            command,
            stdout=subprocess.PIPE,
            stderr=subprocess.STDOUT,
            universal_newlines=True,
            timeout=3600
        )
    except subprocess.TimeoutExpired:
        print(f'Timeout running command for {filename}')
        results.append({'Query': filename, 'Search Time': 'Timeout', 'Result Count': 'Timeout'})
        continue
    except subprocess.CalledProcessError as e:
        print(f'Error running command for {filename}: {e}')
        results.append({'Query': filename, 'Search Time': 'Error', 'Result Count': 'Error'})
        continue
    except KeyboardInterrupt:
        print("Execution interrupted by user")
        break

    search_time_match = re.search(r'Search time \(seconds\): ([\d\.]+)', output.stdout)
    result_count_match = re.search(r'Result count: (\d+)', output.stdout)

    if search_time_match:
        search_time = search_time_match.group(1)
        result_count = result_count_match.group(1) if result_count_match else 'Not found'
        print(f'Search time for {filename}: {search_time}, Result count: {result_count}')
    else:
        search_time = 'No search time found'
        result_count = 'No result count found'
        print(f'No "Search time" found for {filename}')

    results.append({
        'Query': filename,
        'Search Time': search_time,
        'Result Count': result_count
    })

results_df = pd.DataFrame(results)

if os.path.exists(output_file):
    book = load_workbook(output_file)
    with pd.ExcelWriter(output_file, engine='openpyxl', mode='a') as writer:
        writer.book = book
        writer.sheets = {ws.title: ws for ws in book.worksheets}
        results_df.to_excel(writer, index=False, startrow=writer.sheets['Sheet1'].max_row, header=False)
else:
    results_df.to_excel(output_file, index=False)
